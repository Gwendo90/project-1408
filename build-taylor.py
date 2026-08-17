#!/usr/bin/env python3
"""
Baut taylor-songs.json aus Taylor_Swift_Spiel.xlsx.

    python3 build-taylor.py                     # Standardpfade, schreibt die Datei
    python3 build-taylor.py --nur-pruefen       # nichts schreiben, nur Tabelle + Tests
    python3 build-taylor.py --xlsx PFAD

Warum dieses Skript im Repo liegt und nicht im Desktop-Ordner: `build_flags.py` und
`build_songs_json.py` liegen dort, lesen von Pfaden, die es nicht mehr gibt, und schreiben in
die alte Website-Kopie. Beide erzeugen den heutigen Stand nicht wieder (siehe README). Dieses
Skript liest die Excel dort, wo sie liegt, und schreibt neben die Datei, die es erzeugt.

────────────────────────────────────────────────────────────────────────────────────────────
Die Zeitleiste hängt am ERSCHEINUNGSDATUM DES ALBUMS, nicht am Jahr.

258 Songs verteilen sich auf nur 19 Jahrgänge, allein 33 auf 2020 und 31 auf 2024. Nach Jahr
sortiert wäre in einer Zehnerrunde regelmäßig jede Position richtig — das Spiel liefe leer.
Mit dem Albumdatum werden daraus 16 klar geordnete Ären; Gleichstand gibt es nur noch
innerhalb eines Albums, und dort entscheidet die Tracknummer.

Das Datum steht deshalb hier im Skript und nicht in der Excel: es gilt je Album, nicht je
Song. 40 Zeilen statt 258, und Korrekturen bleiben an einer Stelle.
────────────────────────────────────────────────────────────────────────────────────────────

IDs sind eingefroren. `001`–`258` folgen der Zeilenreihenfolge der Excel. Sie stehen in der
Statistik (`tipps.song_id`) und dürfen sich nie verschieben, sonst zeigt eine alte Zeile auf
einen anderen Song – dieselbe Regel wie bei den ESC-IDs. Das Skript vergleicht deshalb gegen
eine vorhandene taylor-songs.json und bricht bei Abweichung ab. Neue Songs ANHÄNGEN, niemals
einsortieren.
"""

import argparse
import json
import re
import sys
from collections import OrderedDict, Counter
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path.home() / "Desktop/project-1408/Links Online Spiel/Taylor_Swift_Spiel.xlsx"
SHEET = "Songs"
ZIEL = Path(__file__).parent / "taylor-songs.json"


# ── Ären ────────────────────────────────────────────────────────────────────────────────────
# Der Slug verbindet Song und Kartendesign. Die zwölf Vorderseiten aus
# "taylors version/Taylors version/" tragen genau diese Farben:
#
#   debut      #a4caa4   151_Front.svg        salbei
#   fearless   #f0c080   151_Front (2).svg    gold
#   speaknow   #c8a9ca   151_Front (3).svg    lila
#   red        #7b2e3a   151_Front (4).svg    dunkelrot
#   1989       #b4e5f9   151_Front (5).svg    hellblau
#   reputation #000000   151_Front (6).svg    schwarz, weisse Schrift
#   lover      #f6b0cc   151_Front (7).svg    rosa
#   folklore   #cccac0   151_Front (8).svg    grau
#   evermore   #c5ac8f   151_Front (9).svg    taupe
#   midnights  #242d48   151_Front (10).svg   navy, weisse Schrift
#   ttpd       #ffffff   151_Front (11).svg   weiss, schwarze Schrift
#   showgirl   #e1581e   151_Front (12).svg   orange
#
# DREI GRUPPEN HABEN KEIN DESIGN – zusammen 16 Songs:
#   film       12 Songs aus 10 Filmen (🎬)
#   weihnacht   3 Songs (🎄)
#   ronan       1 Song  (🕯️, "Ronan")
# Bis eine Vorderseite dafür vorliegt, braucht der Client eine Ersatzfarbe.
ERA_VON_EMOJI = {
    "🤠": "debut",
    "💛": "fearless",
    "💜": "speaknow",
    "❤️": "red",
    "🩵": "1989",
    "🐍": "reputation",
    "💗": "lover",
    "🌲": "folklore",
    "🍂": "evermore",
    "🌌": "midnights",
    "🤍": "ttpd",
    "🧡": "showgirl",
    "🎬": "film",
    "🎄": "weihnacht",
    "🕯️": "ronan",
}


# ── Erscheinungsdaten je Album ──────────────────────────────────────────────────────────────
# Schlüssel ist (Jahr, Album) und nicht der Albumname allein: "Einzelsingle" kommt dreimal vor
# (2012, 2019, 2023) und "The Hunger Games" zweimal (2011, 2012).
#
# Bei Deluxe-Ausgaben, die am selben Tag wie das Album erschienen (Speak Now, Red, 1989,
# Midnights 3am), steht bewusst dasselbe Datum – innerhalb eines Tages entscheidet die
# Tracknummer, und die Ausgaben gehören zusammen.
#
# Bei Soundtracks und Singles zählt der Tag, an dem DER SONG erschien, nicht der des Films
# oder des Sammelalbums. Für ein Zeitleistenspiel ist das der richtige Anker.
#
# ⚠ GEPRÜFT WERDEN SOLLTEN die mit "?" markierten Zeilen. Alles andere sind gut belegte
# Veröffentlichungstage.
DATUM = {
    ("2006", "Taylor Swift"):                        "2006-10-24",
    ("2007", "Taylor Swift (Deluxe)"):               "2007-11-06",  # ? Deluxe-Ausgabe des Debüts
    ("2007", "The Taylor Swift Holiday Collection"): "2007-10-14",
    ("2008", "Fearless"):                            "2008-11-11",
    ("2009", "Fearless (Platinum Edition)"):         "2009-10-26",
    ("2009", "Hannah Montana: The Movie"):           "2009-03-24",
    ("2010", "Speak Now"):                           "2010-10-25",
    ("2010", "Speak Now (Deluxe)"):                  "2010-10-25",
    ("2010", "Valentine's Day"):                     "2010-01-19",  # Single vor dem Soundtrack
    ("2011", "The Hunger Games"):                    "2011-12-26",  # Safe & Sound
    ("2012", "Einzelsingle"):                        "2012-09-08",  # Ronan
    ("2012", "Red"):                                 "2012-10-22",
    ("2012", "Red (Deluxe)"):                        "2012-10-22",
    ("2012", "The Hunger Games"):                    "2012-03-27",  # Eyes Open
    ("2013", "One Chance"):                          "2013-10-21",  # Sweeter Than Fiction
    ("2014", "1989"):                                "2014-10-27",
    ("2014", "1989 (Deluxe)"):                       "2014-10-27",
    ("2016", "Fifty Shades Darker"):                 "2016-12-09",
    ("2017", "reputation"):                          "2017-11-10",
    ("2019", "Cats"):                                "2019-12-20",  # Sammelalbum, s. DATUM_SONG
    ("2019", "Einzelsingle"):                        "2019-12-06",  # Christmas Tree Farm
    ("2019", "Lover"):                               "2019-08-23",
    ("2020", "Miss Americana"):                      "2020-01-31",  # Only the Young
    ("2020", "folklore"):                            "2020-07-24",
    ("2020", "evermore"):                            "2020-12-11",
    ("2021", "evermore (Deluxe)"):                   "2021-01-07",
    ("2021", "Fearless (Taylor's Version)"):         "2021-04-09",
    ("2021", "Spirit Untamed"):                      "2021-09-17",  # Wildest Dreams (TV)
    ("2021", "Red (Taylor's Version)"):              "2021-11-12",
    ("2022", "Where the Crawdads Sing"):             "2022-06-24",  # Carolina
    ("2022", "Midnights"):                           "2022-10-21",
    ("2022", "Midnights (3am Edition)"):             "2022-10-21",
    ("2022", "Midnights (Bonus)"):                   "2022-10-21",  # Hits Different
    ("2023", "Einzelsingle"):                        "2023-05-26",  # You're Losing Me
    ("2023", "Speak Now (Taylor's Version)"):        "2023-07-07",
    ("2023", "1989 (Taylor's Version)"):             "2023-10-27",
    ("2024", "The Tortured Poets Department"):       "2024-04-19",
    ("2024", "TTPD: The Anthology"):                 "2024-04-19",
    ("2025", "The Life of a Showgirl"):              "2025-10-03",
    ("2026", "Toy Story 5"):                         "2026-06-19",  # ? Kinostart, Song unbekannt
}

# Einzelne Songs, die früher erschienen als das Album, unter dem sie in der Excel stehen.
# "Beautiful Ghosts" war einen Monat vor dem Cats-Soundtrack draussen; ohne diese Ausnahme
# stünde die Karte fünf Wochen zu spät in der Zeitleiste.
DATUM_SONG = {
    "109": "2019-11-15",  # Beautiful Ghosts
}


def sid_aus(url):
    """Die Track-ID aus dem Apple-Music-Link (`?i=…`). Ohne sie kein Rückweg zum Store."""
    m = re.search(r"[?&]i=(\d+)", str(url or ""))
    return m.group(1) if m else None


def lies_excel(pfad):
    wb = load_workbook(pfad, read_only=True, data_only=True)
    ws = wb[SHEET]
    kopf = [str(c or "").strip() for c in next(ws.iter_rows(max_row=1, values_only=True))]
    erwartet = ["Jahr", "Emoji", "Album", "Interpret", "Songtitel", "Nr.",
                "Apple Music ✓", "Extented Preview", "Kategorie"]
    if kopf[:len(erwartet)] != erwartet:
        sys.exit(f"Kopfzeile passt nicht.\n  erwartet: {erwartet}\n  gefunden: {kopf}")
    zeilen = [r for r in ws.iter_rows(min_row=2, values_only=True)
              if any(c is not None and str(c).strip() for c in r)]
    wb.close()
    return zeilen


def baue(zeilen):
    songs, fehler = OrderedDict(), []
    for i, r in enumerate(zeilen, start=1):
        jahr, emoji, album, interpret, titel, nr, u, pr, kat = (str(x).strip() if x is not None else "" for x in r[:9])
        sid = f"{i:03d}"
        datum = DATUM_SONG.get(sid) or DATUM.get((jahr, album))
        if not datum:
            fehler.append(f"{sid}: kein Datum für ({jahr}, {album!r}) – in DATUM ergänzen")
            datum = ""
        elif datum[:4] != jahr:
            fehler.append(f"{sid}: Datum {datum} widerspricht Jahr {jahr} ({album})")
        era = ERA_VON_EMOJI.get(emoji)
        if not era:
            fehler.append(f"{sid}: unbekanntes Emoji {emoji!r} – in ERA_VON_EMOJI ergänzen")
            era = "film"
        if not pr.startswith("https://audio-ssl.itunes.apple.com/"):
            fehler.append(f"{sid}: Vorschau sieht nicht nach iTunes aus: {pr[:60]}")
        for feld, wert in (("Songtitel", titel), ("Album", album), ("Interpret", interpret)):
            if not wert:
                fehler.append(f"{sid}: {feld} ist leer")
        songs[sid] = {
            "year": jahr,
            "artist": interpret,
            "title": titel,
            "album": album,
            "era": era,
            "emoji": emoji,
            "date": datum,
            "track": int(nr) if str(nr).isdigit() else None,
            "cat": kat,
            "sid": sid_aus(u),
            "u": u,
            "pr": pr,
        }
    return songs, fehler


def pruefe_ids(songs, ziel):
    """Verschobene IDs verfälschen die Statistik rückwirkend. Lieber abbrechen."""
    if not ziel.exists():
        return []
    alt = json.loads(ziel.read_text(encoding="utf-8"))
    fehler = []
    for sid, a in alt.items():
        n = songs.get(sid)
        if n is None:
            fehler.append(f"{sid} fehlt jetzt: {a['title']!r} – IDs nicht neu vergeben")
        elif n["title"] != a["title"]:
            fehler.append(f"{sid} zeigt jetzt auf {n['title']!r}, vorher {a['title']!r}")
    return fehler


def tabelle(songs):
    """Ären und Alben in Spielreihenfolge – die Liste, an der die Daten zu prüfen sind."""
    alben = OrderedDict()
    for s in songs.values():
        alben.setdefault((s["date"], s["year"], s["album"], s["era"], s["emoji"]), 0)
        alben[(s["date"], s["year"], s["album"], s["era"], s["emoji"])] += 1
    print(f"\n{'Datum':<12} {'Ära':<11} {'Album':<40} {'Songs':>5}")
    print("─" * 72)
    for (datum, _jahr, album, era, emoji), n in sorted(alben.items()):
        print(f"{datum:<12} {era:<11} {emoji} {album:<38} {n:>5}")
    print("─" * 72)
    eras = Counter(s["era"] for s in songs.values())
    print(f"{len(songs)} Songs, {len(alben)} Alben, {len(eras)} Ären")
    ohne = {"film", "weihnacht", "ronan"} & set(eras)
    if ohne:
        print("ohne Kartendesign: " + ", ".join(f"{e} ({eras[e]})" for e in sorted(ohne)))
    kurz = [sid for sid, s in songs.items() if ".ep." not in s["pr"]]
    print(f"kurze Vorschauen (30 s statt 90 s): {len(kurz)} – {', '.join(kurz)}")


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--xlsx", type=Path, default=XLSX)
    p.add_argument("--ziel", type=Path, default=ZIEL)
    p.add_argument("--nur-pruefen", action="store_true", help="nichts schreiben")
    a = p.parse_args()

    if not a.xlsx.exists():
        sys.exit(f"Excel nicht gefunden: {a.xlsx}")

    songs, fehler = baue(lies_excel(a.xlsx))
    fehler += pruefe_ids(songs, a.ziel)
    tabelle(songs)

    if fehler:
        print("\nFEHLER:")
        for f in fehler:
            print("  " + f)
        sys.exit(1)
    print("\nalle Prüfungen bestanden")

    if a.nur_pruefen:
        return
    a.ziel.write_text(json.dumps(songs, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    print(f"geschrieben: {a.ziel} ({a.ziel.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
